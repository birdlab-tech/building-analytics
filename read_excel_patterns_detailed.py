"""
Read Excel filter patterns in detail, checking both formula and data modes
"""
import openpyxl

filepath = r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Heating-Cooling.xlsx"

print("="*80)
print("READING Bs1 SHEET - Filter Pattern Columns")
print("="*80)

# Read with formulas
wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
ws_formulas = wb_formulas['Bs1']

# Read with data
wb_data = openpyxl.load_workbook(filepath, data_only=True)
ws_data = wb_data['Bs1']

print("\nColumn Headers and first few rows:")
print("-"*80)

# Check columns AH (34), AI (35), AJ (36), AK (37)
columns_to_check = {
    'AH': 34,
    'AI': 35,
    'AJ': 36,
    'AK': 37
}

for col_name, col_num in columns_to_check.items():
    print(f"\n=== Column {col_name} (column {col_num}) ===")

    for row in range(1, 11):
        cell_formula = ws_formulas.cell(row=row, column=col_num).value
        cell_data = ws_data.cell(row=row, column=col_num).value

        # Check if it's a formula
        if cell_formula and isinstance(cell_formula, str) and cell_formula.startswith('='):
            print(f"Row {row} [FORMULA]: {cell_formula}")
        else:
            # Show raw data value and its type
            if cell_data is not None:
                print(f"Row {row} [DATA]: {repr(cell_data)} (type: {type(cell_data).__name__})")
            else:
                print(f"Row {row}: (empty)")

# Also check the Menus sheet if it exists (might contain the dropdown options)
if 'Menus' in wb_data.sheetnames:
    print("\n" + "="*80)
    print("MENUS SHEET (may contain dropdown options)")
    print("="*80)
    ws_menus = wb_data['Menus']
    for row in range(1, min(21, ws_menus.max_row + 1)):
        values = []
        for col in range(1, min(6, ws_menus.max_column + 1)):
            val = ws_menus.cell(row=row, column=col).value
            if val:
                values.append(str(val))
        if values:
            print(f"Row {row}: {' | '.join(values)}")
