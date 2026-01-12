"""
Scan all columns in Bs1 sheet to find where search patterns are stored
"""
import openpyxl

filepath = r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Heating-Cooling.xlsx"

wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb['Bs1']

print("="*80)
print("Scanning Bs1 sheet for text patterns (rows 2-15, looking for searchable text)")
print("="*80)

# Scan columns from A to AZ (first 52 columns)
for col_num in range(1, 53):
    col_letter = openpyxl.utils.get_column_letter(col_num)

    # Get values from rows 2-15 to see if there's interesting text
    values = []
    for row in range(2, 16):
        val = ws.cell(row=row, column=col_num).value
        if val and isinstance(val, str) and len(val) > 1 and not val.startswith('='):
            values.append(val)

    # If we found interesting text, print it
    if values and any(len(v) > 2 for v in values):
        print(f"\nColumn {col_letter} (#{col_num}):")
        for i, val in enumerate(values[:10], start=2):
            if len(val) > 2:  # Only show non-trivial strings
                print(f"  Row {i}: {val[:60]}")

print("\n" + "="*80)
print("Checking row 2 across all columns for search patterns:")
print("="*80)

# Check row 2 specifically
for col_num in range(1, 53):
    col_letter = openpyxl.utils.get_column_letter(col_num)
    val = ws.cell(row=2, column=col_num).value

    if val and isinstance(val, str) and len(val) > 2:
        print(f"Column {col_letter}: {val[:80]}")
