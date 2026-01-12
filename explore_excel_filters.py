"""
Explore the Excel label filtering system structure
"""
import openpyxl
import pandas as pd
from pathlib import Path

# File paths
excel_files = [
    r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Heating-Cooling.xlsx",
    r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Lighting.xlsx",
    r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes - AHUs).xlsx"
]

def explore_excel_structure(filepath):
    """Explore the structure of an Excel filtering workbook"""
    print(f"\n{'='*80}")
    print(f"FILE: {Path(filepath).name}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    print(f"\nSheets: {wb.sheetnames}")

    # Explore Labels sheet
    if 'Labels' in wb.sheetnames:
        ws = wb['Labels']
        print(f"\n--- LABELS SHEET ---")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols")
        print("\nFirst 3 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            print(f"  Row {i}: {row[:15]}")

    # Explore Bs1 (Blocker 1)
    if 'Bs1' in wb.sheetnames:
        ws = wb['Bs1']
        print(f"\n--- Bs1 SHEET (Blockers 1) ---")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols")
        print("\nFirst 3 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            print(f"  Row {i}: {row[:15]}")

        # Check column AK (column 37)
        print("\nColumn AK (col 37) sample values:")
        for i in range(1, min(6, ws.max_row+1)):
            cell_value = ws.cell(row=i, column=37).value
            print(f"  Row {i}: {cell_value}")

    # Explore Ts (Targets)
    if 'Ts' in wb.sheetnames:
        ws = wb['Ts']
        print(f"\n--- Ts SHEET (Targets) ---")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols")
        print("\nFirst 3 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            print(f"  Row {i}: {row[:15]}")

    # Check for merged cells in Bs-Ts sheet
    if 'Bs-Ts' in wb.sheetnames:
        ws = wb['Bs-Ts']
        print(f"\n--- Bs-Ts SHEET ---")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols")
        print(f"Merged cells: {list(ws.merged_cells)[:5]}")  # Show first 5 merged cells
        print("\nFirst 3 rows:")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            print(f"  Row {i}: {row[:10]}")

# Explore first file in detail
explore_excel_structure(excel_files[0])

print("\n" + "="*80)
print("Summary of all three files:")
print("="*80)
for filepath in excel_files:
    wb = openpyxl.load_workbook(filepath, data_only=False)
    labels_count = wb['Labels'].max_row if 'Labels' in wb.sheetnames else 0
    print(f"\n{Path(filepath).stem}:")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Total labels: {labels_count}")
