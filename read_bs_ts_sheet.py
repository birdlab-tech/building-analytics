"""
Read the Bs-Ts sheet which consolidates filter criteria
"""
import openpyxl

filepath = r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Heating-Cooling.xlsx"

wb = openpyxl.load_workbook(filepath, data_only=True)

if 'Bs-Ts' in wb.sheetnames:
    ws = wb['Bs-Ts']
    print("="*80)
    print("Bs-Ts SHEET (Consolidated Filter Criteria)")
    print("="*80)
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols\n")

    # Print all rows
    for row_num in range(1, min(32, ws.max_row + 1)):
        row_data = []
        for col_num in range(1, ws.max_column + 1):
            val = ws.cell(row=row_num, column=col_num).value
            if val is not None:
                # Convert to string and truncate if too long
                val_str = str(val)[:50]
                row_data.append(val_str)
            else:
                row_data.append('')

        # Only print if row has data
        if any(row_data):
            print(f"Row {row_num:2d}: {' | '.join(row_data)}")

print("\n" + "="*80)
print("Checking columns with formulas in Bs1 to understand the logic")
print("="*80)

# Let's also check if there's a specific input area
ws_bs1 = wb['Bs1']
print("\nRows 1-3, columns 30-40 (around where filter criteria might be):")
for row in range(1, 4):
    print(f"\nRow {row}:")
    for col in range(30, 41):
        val = ws_bs1.cell(row=row, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        if val is not None:
            print(f"  {col_letter}{row}: {repr(val)}")
