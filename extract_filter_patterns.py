"""
Extract filter patterns from Excel label filtering workbooks
"""
import openpyxl
import json
from pathlib import Path

def extract_filter_config(filepath, system_name):
    """Extract filter configuration from an Excel workbook"""
    print(f"\nProcessing: {system_name}")
    print("="*60)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    config = {
        'system': system_name,
        'labels': [],
        'blockers': [],
        'targets': []
    }

    # Extract labels
    if 'Labels' in wb.sheetnames:
        ws = wb['Labels']
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if label:
                config['labels'].append(label)
        print(f"Labels found: {len(config['labels'])}")

    # Extract blocker patterns (Bs1, Bs2, Bs3, Bs4)
    for blocker_num in range(1, 5):
        sheet_name = f'Bs{blocker_num}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            blocker_config = {
                'name': sheet_name,
                'patterns': []
            }

            # Look for pattern columns (typically AH, AI, AJ)
            # AH = column 34 (pattern), AI = column 35 (fail value), AJ = column 36 (pass value)
            for row in range(2, 32):  # Check up to row 31 for patterns
                pattern = ws.cell(row=row, column=34).value  # Column AH
                if pattern:
                    blocker_config['patterns'].append({
                        'pattern': pattern,
                        'action': 'block'  # Default action is to block/exclude
                    })

            if blocker_config['patterns']:
                config['blockers'].append(blocker_config)
                print(f"{sheet_name}: {len(blocker_config['patterns'])} patterns")

    # Extract target patterns (Ts sheet)
    if 'Ts' in wb.sheetnames:
        ws = wb['Ts']
        target_config = {
            'name': 'Ts',
            'patterns': []
        }

        # For Ts sheet, patterns are in column AI (column 35)
        for row in range(2, 32):
            pattern = ws.cell(row=row, column=35).value  # Column AI
            if pattern:
                target_config['patterns'].append({
                    'pattern': pattern,
                    'action': 'include'  # Targets include labels
                })

        if target_config['patterns']:
            config['targets'].append(target_config)
            print(f"Ts: {len(target_config['patterns'])} patterns")

    return config

# Process all three Excel files
excel_files = {
    'heating_cooling': r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Heating-Cooling.xlsx",
    'lighting': r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes) Lighting.xlsx",
    'ahus': r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\Label Filters (Las Mercedes - AHUs).xlsx"
}

all_configs = {}

for system_name, filepath in excel_files.items():
    try:
        config = extract_filter_config(filepath, system_name)
        all_configs[system_name] = config
    except Exception as e:
        print(f"Error processing {system_name}: {e}")

# Save to JSON
output_file = r"C:\Users\ahami\OneDrive\Documents\KCL PhD\ResearchProposal\Independent Building Analytics\label_filter_configs.json"
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_configs, f, indent=2, ensure_ascii=False)

print("\n" + "="*60)
print(f"Filter configurations saved to: {Path(output_file).name}")
print("\nSummary:")
for system_name, config in all_configs.items():
    print(f"\n{system_name.upper()}:")
    print(f"  Labels: {len(config['labels'])}")
    print(f"  Blockers: {len(config['blockers'])}")
    print(f"  Targets: {len(config['targets'])}")

    # Show sample patterns
    if config['blockers']:
        print(f"  Sample blocker patterns:")
        for blocker in config['blockers'][:2]:
            for pattern in blocker['patterns'][:3]:
                print(f"    - {pattern['pattern']}")
